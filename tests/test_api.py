import asyncio
import io

from fastapi import UploadFile
from openpyxl import load_workbook

from coad_parser.api import _criar_planilha, _processar_um_arquivo
from coad_parser.schemas.document_data import DocumentData


def processar(nome: str, conteudo: str) -> dict:
    arquivo = UploadFile(filename=nome, file=io.BytesIO(conteudo.encode("utf-8")))
    return asyncio.run(_processar_um_arquivo(arquivo))


def test_rejeita_html_vazio():
    resultado = processar("vazio.html", "   ")

    assert resultado["sucesso"] is False
    assert resultado["erro"] == "Arquivo HTML vazio"


def test_rejeita_pagina_de_login():
    resultado = processar(
        "login.html",
        '<html><title>Login</title><form><input type="password"></form></html>',
    )

    assert resultado["sucesso"] is False
    assert "login/autenticação" in resultado["erro"]


def test_rejeita_html_sem_campos_esperados():
    resultado = processar(
        "pagina.html",
        "<html><body><p>Conteúdo sem relação com um documento COAD.</p></body></html>",
    )

    assert resultado["sucesso"] is False
    assert "sem os campos esperados" in resultado["erro"]


def test_aceita_documento_com_campo_esperado():
    resultado = processar(
        "sei_123456.html",
        "<html><body><p>Autorização Direta Nº 12/2025</p></body></html>",
    )

    assert resultado["sucesso"] is True
    assert resultado["dados"]["numero_autorizacao"] == "12/2025"


def test_planilha_separa_completos_dados_faltantes_e_htmls_quebrados():
    dados_completos = {campo: "valor" for campo in DocumentData.model_fields}
    dados_completos["ucs_envolvidas"] = "PARNA da Tijuca"
    dados_completos["coordenadas_brutas"] = None
    resposta = _criar_planilha(
        [
            {
                "arquivo": "completo.html",
                "sucesso": True,
                "dados": dados_completos,
            },
            {
                "arquivo": "incompleto.html",
                "sucesso": True,
                "dados": {"numero_sei": "123456"},
            },
            {
                "arquivo": "quebrado.html",
                "sucesso": False,
                "erro": "Arquivo HTML vazio",
            },
        ]
    )
    corpo = asyncio.run(_ler_resposta(resposta))
    planilha = load_workbook(io.BytesIO(corpo))

    assert planilha.sheetnames == [
        "Registros completos",
        "Dados faltantes",
        "HTMLs quebrados",
    ]
    assert planilha["Registros completos"].max_column == 11
    assert planilha["Dados faltantes"].max_column == 11
    assert "arquivo_origem" not in {
        celula.value for celula in planilha["Registros completos"][1]
    }
    assert "arquivo_origem" not in {
        celula.value for celula in planilha["Dados faltantes"][1]
    }
    assert planilha["Registros completos"]["A1"].value == planilha["Dados faltantes"]["A1"].value
    assert planilha["Dados faltantes"]["A2"].value is None
    assert planilha["HTMLs quebrados"]["A2"].value == "quebrado.html"
    assert planilha["HTMLs quebrados"]["B2"].value == "Arquivo HTML vazio"
    assert planilha["HTMLs quebrados"]["A1"].value == "arquivo"


def test_planilha_ordena_numero_autorizacao_por_ano_e_numero():
    resultados = []
    for numero in ("10/2025", "2/2025", "8/2024"):
        dados = {campo: "valor" for campo in DocumentData.model_fields}
        dados["numero_autorizacao"] = numero
        dados["ucs_envolvidas"] = "PARNA da Tijuca"
        dados["coordenadas_brutas"] = None
        resultados.append(
            {
                "arquivo": f"{numero.replace('/', '-')}.html",
                "sucesso": True,
                "dados": dados,
            }
        )

    resposta = _criar_planilha(resultados)
    corpo = asyncio.run(_ler_resposta(resposta))
    planilha = load_workbook(io.BytesIO(corpo))

    valores = [
        celula.value
        for celula in planilha["Registros completos"]["A"][1:]
    ]
    assert valores == ["8/2024", "2/2025", "10/2025"]


def test_planilha_coloca_numero_autorizacao_vazio_no_final():
    resultados = [
        {
            "arquivo": "sem-numero.html",
            "sucesso": True,
            "dados": {"numero_sei": "123456"},
        },
        {
            "arquivo": "com-numero.html",
            "sucesso": True,
            "dados": {"numero_autorizacao": "2/2025"},
        },
    ]

    resposta = _criar_planilha(resultados)
    corpo = asyncio.run(_ler_resposta(resposta))
    planilha = load_workbook(io.BytesIO(corpo))

    valores = [
        celula.value
        for celula in planilha["Dados faltantes"]["A"][1:]
    ]
    assert valores == ["2/2025", None]


async def _ler_resposta(resposta) -> bytes:
    partes = [parte async for parte in resposta.body_iterator]
    return b"".join(partes)
